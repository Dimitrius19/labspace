"""
Build 37 5-year financial models (.xlsx) for the Greek govtech PASS shortlist.

Each model has 6 sheets — Assumptions, Revenue, Headcount_Costs, P_and_L,
Cash_Flow, KPIs — with live Excel formulas wiring Assumptions through to KPIs.
Run from repo root.
"""

from __future__ import annotations
import os
from dataclasses import dataclass, field
from typing import List, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_DIR = "/Users/dimitristryfon/Desktop/labspace/src/data/govtech/financials"
os.makedirs(OUT_DIR, exist_ok=True)

# Cell colours
INPUT_FILL = PatternFill("solid", fgColor="FFF9C4")    # light yellow
TOTAL_FILL = PatternFill("solid", fgColor="E3F2FD")    # light blue
HEADER_FILL = PatternFill("solid", fgColor="263238")    # dark slate
SECTION_FILL = PatternFill("solid", fgColor="ECEFF1")
WHITE = PatternFill("solid", fgColor="FFFFFF")

BOLD = Font(bold=True)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
SECTION_FONT = Font(bold=True, color="263238", size=11)
ITAL = Font(italic=True, color="546E7A")

THIN = Side(style="thin", color="CFD8DC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

EUR = "€#,##0"
EUR_DEC = "€#,##0.00"
PCT = "0.0%"
INT = "#,##0"


@dataclass
class Drivers:
    """Per-idea financial drivers. €. All numbers are absolutes (not in '000)."""
    company: str
    ministry: str
    buyer: str
    # Pricing
    tier1_price: int              # pilot / Tier-1 contract value
    tier2_acv: int                # year-one ACV per Tier-2 customer
    tier3_acv: int                # framework / national ACV
    seats_per_t2: int = 30        # avg seats per Tier-2 customer
    seat_price: int = 3000        # per-seat blended €/yr where applicable
    # Adoption — counts per year [Y1..Y5]
    t1_count: Tuple[int, int, int, int, int] = (1, 2, 1, 0, 0)
    t2_count: Tuple[int, int, int, int, int] = (0, 1, 3, 5, 7)
    t3_count: Tuple[int, int, int, int, int] = (0, 0, 0, 1, 2)
    # Headcount build [Y1..Y5]
    founders: Tuple[int, int, int, int, int] = (3, 3, 3, 3, 3)
    engineering: Tuple[int, int, int, int, int] = (1, 3, 5, 8, 11)
    product: Tuple[int, int, int, int, int] = (0, 1, 1, 2, 2)
    sales: Tuple[int, int, int, int, int] = (0, 1, 2, 3, 4)
    cs: Tuple[int, int, int, int, int] = (0, 1, 2, 3, 4)
    ga: Tuple[int, int, int, int, int] = (0, 1, 1, 2, 3)
    fte_cost: int = 72000         # Athens loaded €
    # COGS
    hosting_per_customer: int = 8000
    cs_pct_revenue: float = 0.05  # CS additional COGS %
    # Opex ratios (% of revenue)
    rd_pct: float = 0.0           # R&D drawn from headcount already; extra non-FTE R&D
    sm_pct: float = 0.10          # S&M non-personnel %
    ga_pct: float = 0.06          # G&A non-personnel %
    # Capital
    pre_seed: int = 700000
    seed: int = 3000000
    series_a: int = 8000000
    pre_seed_year: int = 1
    seed_year: int = 2
    series_a_year: int = 4
    # Targets
    gross_margin_target: float = 0.72


# ---- Per-idea drivers (37 entries) ----------------------------------------
# Numbers come from the Business Plan sections of the regenerated memo packs.
# Where the memo gives a range, mid-point is used.

DRIVERS: dict[str, Drivers] = {
    # ===== Original 12 =====
    "tourism-ama-cross-check-agent": Drivers(
        "AMA Cross-Check", "tourism", "Ministry of Tourism / MΗΤΕ",
        tier1_price=55_000, tier2_acv=280_000, tier3_acv=1_100_000,
        t1_count=(2, 2, 1, 0, 0), t2_count=(0, 2, 4, 6, 8), t3_count=(0, 0, 0, 1, 2),
        pre_seed=600_000, seed=2_800_000, series_a=8_000_000,
    ),
    "digital-gov-citizen-concierge": Drivers(
        "Citizen Concierge", "digital-governance", "ΓΓΠΣΨΔ",
        tier1_price=80_000, tier2_acv=400_000, tier3_acv=1_800_000,
        seats_per_t2=200, seat_price=400,
        t1_count=(2, 2, 1, 0, 0), t2_count=(0, 2, 5, 9, 13), t3_count=(0, 0, 1, 1, 2),
        pre_seed=900_000, seed=3_500_000, series_a=10_000_000,
        engineering=(2, 4, 7, 10, 13), gross_margin_target=0.75,
    ),
    "archaeology-permit-triage-ephorate": Drivers(
        "Ephorate Triage", "culture", "Ministry of Culture / ΥΠΠΟ",
        tier1_price=50_000, tier2_acv=220_000, tier3_acv=700_000,
        t1_count=(2, 3, 2, 1, 0), t2_count=(0, 2, 5, 8, 12), t3_count=(0, 0, 0, 1, 1),
        pre_seed=600_000, seed=2_500_000, series_a=6_000_000,
    ),
    "energy-res-permit-crosschecker": Drivers(
        "RES Permit Cross-Checker", "environment-energy", "ΥΠΕΝ / ΡΑΑΕΥ",
        tier1_price=90_000, tier2_acv=450_000, tier3_acv=1_400_000,
        t1_count=(2, 3, 2, 0, 0), t2_count=(0, 2, 5, 9, 14), t3_count=(0, 0, 1, 2, 3),
        pre_seed=800_000, seed=3_500_000, series_a=10_000_000,
        engineering=(2, 4, 7, 10, 14),
    ),
    "rrf-milestone-tracker": Drivers(
        "RRF Milestone Tracker", "state-cross-cutting", "ΓΓΠΘ / Coordinating Ministry",
        tier1_price=100_000, tier2_acv=500_000, tier3_acv=1_500_000,
        t1_count=(2, 2, 1, 0, 0), t2_count=(0, 2, 4, 7, 10), t3_count=(0, 0, 1, 2, 3),
        pre_seed=800_000, seed=3_000_000, series_a=9_000_000,
    ),
    "espa-screening-copilot": Drivers(
        "ESPA Screening", "development", "Ministry of National Economy / ΕΥΣΕ",
        tier1_price=80_000, tier2_acv=400_000, tier3_acv=1_300_000,
        t1_count=(2, 3, 2, 0, 0), t2_count=(0, 2, 5, 9, 13), t3_count=(0, 0, 1, 1, 2),
        pre_seed=750_000, seed=3_000_000, series_a=8_500_000,
    ),
    "efka-pension-calc-copilot": Drivers(
        "Sintaxis Copilot", "labour-social-security", "EFKA / ΥΠΕΚΥΠ",
        tier1_price=52_000, tier2_acv=375_000, tier3_acv=2_000_000,
        seats_per_t2=125, seat_price=3000,
        t1_count=(1, 2, 1, 0, 0), t2_count=(0, 1, 3, 5, 7), t3_count=(0, 0, 0, 1, 2),
        pre_seed=700_000, seed=3_000_000, series_a=10_000_000,
        engineering=(2, 4, 6, 9, 12), gross_margin_target=0.72,
    ),
    "aade-elenxis-auditor-copilot": Drivers(
        "ELENXIS Auditor Copilot", "finance-aade", "AADE Governor",
        tier1_price=90_000, tier2_acv=500_000, tier3_acv=1_800_000,
        seats_per_t2=70, seat_price=7200,
        t1_count=(2, 2, 1, 0, 0), t2_count=(0, 2, 4, 7, 10), t3_count=(0, 0, 1, 1, 2),
        pre_seed=900_000, seed=3_800_000, series_a=10_000_000,
    ),
    "mfa-eastmed-osint-monitor": Drivers(
        "EastMed OSINT", "foreign-affairs", "MFA / Hellenic National Defence GS",
        tier1_price=85_000, tier2_acv=350_000, tier3_acv=1_200_000,
        t1_count=(2, 2, 1, 0, 0), t2_count=(0, 1, 3, 5, 7), t3_count=(0, 0, 0, 1, 2),
        pre_seed=800_000, seed=3_000_000, series_a=8_000_000,
    ),
    "maritime-barren-line-subsidy-optimiser": Drivers(
        "Barren-Line Optimiser", "maritime-insular", "ΥΝΑΝΠ",
        tier1_price=60_000, tier2_acv=280_000, tier3_acv=900_000,
        t1_count=(2, 2, 1, 0, 0), t2_count=(0, 1, 2, 4, 6), t3_count=(0, 0, 0, 1, 1),
        pre_seed=550_000, seed=2_400_000, series_a=6_000_000,
    ),
    "migration-coi-drafting-copilot": Drivers(
        "COI Drafting Copilot", "migration-asylum", "Greek Asylum Service / ΥΜΑ",
        tier1_price=70_000, tier2_acv=320_000, tier3_acv=1_100_000,
        seats_per_t2=80, seat_price=4000,
        t1_count=(2, 2, 1, 0, 0), t2_count=(0, 1, 3, 5, 7), t3_count=(0, 0, 0, 1, 1),
        pre_seed=700_000, seed=2_800_000, series_a=7_500_000,
    ),
    "digital-gov-llm-eval-redteam": Drivers(
        "LLM Eval & Red-Team", "digital-governance", "ΓΓΠΣΨΔ + cross-ministry",
        tier1_price=90_000, tier2_acv=450_000, tier3_acv=1_600_000,
        t1_count=(3, 4, 3, 1, 0), t2_count=(0, 3, 6, 10, 15), t3_count=(0, 0, 1, 2, 4),
        pre_seed=900_000, seed=4_000_000, series_a=12_000_000,
        engineering=(3, 5, 8, 12, 16), gross_margin_target=0.78,
    ),

    # ===== 25 new PASS shortlist =====
    "social-opeka-benefit-claim-copilot": Drivers(
        "Pronoia Copilot", "social-cohesion-family", "OPA / ΥΕΚΑ",
        tier1_price=55_000, tier2_acv=380_000, tier3_acv=1_350_000,
        seats_per_t2=130, seat_price=2900,
        t1_count=(2, 2, 1, 0, 0), t2_count=(0, 2, 4, 6, 9), t3_count=(0, 0, 1, 1, 2),
        pre_seed=750_000, seed=3_200_000, series_a=9_000_000,
        engineering=(2, 4, 6, 9, 12),
    ),
    "civil-protection-post-incident-drafting": Drivers(
        "Μετά-Συμβάντος", "climate-civil-protection", "ΓΓΠΠ General Secretary",
        tier1_price=90_000, tier2_acv=300_000, tier3_acv=700_000,
        t1_count=(2, 2, 1, 0, 0), t2_count=(0, 1, 3, 4, 5), t3_count=(0, 0, 1, 1, 2),
        pre_seed=600_000, seed=2_500_000, series_a=6_000_000,
    ),
    "infrastructure-contractor-progress-monitor": Drivers(
        "Contractor Sentinel", "infrastructure-transport", "ΥΠΥΜΕ / GS Υποδομών",
        tier1_price=120_000, tier2_acv=600_000, tier3_acv=2_000_000,
        t1_count=(2, 2, 1, 0, 0), t2_count=(0, 1, 3, 5, 8), t3_count=(0, 0, 1, 2, 3),
        pre_seed=900_000, seed=3_800_000, series_a=10_000_000,
    ),
    "interior-kede-shared-services-stack": Drivers(
        "Kallikratis Stack", "interior", "ΚΕΔΕ / ΥΠΕΣ",
        tier1_price=180_000, tier2_acv=350_000, tier3_acv=1_500_000,
        seats_per_t2=12, seat_price=29000,  # per-municipality
        t1_count=(2, 3, 2, 0, 0), t2_count=(0, 8, 18, 35, 60), t3_count=(0, 0, 1, 2, 3),
        pre_seed=850_000, seed=3_500_000, series_a=10_000_000,
        engineering=(2, 5, 9, 14, 19),
    ),
    "civil-protection-field-officer-dictation": Drivers(
        "Πυρόπολις", "climate-civil-protection", "ΠΣ / ΓΓΠΠ",
        tier1_price=110_000, tier2_acv=300_000, tier3_acv=1_400_000,
        seats_per_t2=400, seat_price=750,
        t1_count=(2, 2, 1, 0, 0), t2_count=(0, 1, 3, 5, 8), t3_count=(0, 0, 0, 1, 2),
        pre_seed=800_000, seed=3_200_000, series_a=8_500_000,
    ),
    "aade-transfer-pricing-review-agent": Drivers(
        "TP-Review.gr", "finance-aade", "AADE Governor",
        tier1_price=75_000, tier2_acv=216_000, tier3_acv=900_000,
        seats_per_t2=30, seat_price=7200,
        t1_count=(2, 2, 1, 0, 0), t2_count=(0, 2, 4, 7, 10), t3_count=(0, 0, 0, 1, 2),
        pre_seed=750_000, seed=3_000_000, series_a=8_000_000,
    ),
    "transport-driving-test-scheduler": Drivers(
        "DriveSlot", "infrastructure-transport", "Γ.Γ. Μεταφορών",
        tier1_price=60_000, tier2_acv=350_000, tier3_acv=1_400_000,
        t1_count=(2, 2, 1, 0, 0), t2_count=(0, 2, 4, 6, 9), t3_count=(0, 0, 1, 1, 2),
        pre_seed=600_000, seed=2_600_000, series_a=7_500_000,
    ),
    "justice-echr-monitor-agent": Drivers(
        "Strasbourg Watch", "justice", "Νομικό Γραφείο MFA + ΝΣΚ",
        tier1_price=90_000, tier2_acv=230_000, tier3_acv=800_000,
        t1_count=(2, 2, 1, 0, 0), t2_count=(0, 1, 3, 5, 7), t3_count=(0, 0, 1, 1, 2),
        pre_seed=550_000, seed=2_200_000, series_a=5_500_000,
    ),
    "agri-efet-inspection-prioritisation": Drivers(
        "Trofima Sentinel", "rural-development", "EFET Governor",
        tier1_price=55_000, tier2_acv=350_000, tier3_acv=1_200_000,
        seats_per_t2=80, seat_price=3500,
        t1_count=(2, 2, 1, 0, 0), t2_count=(0, 2, 4, 6, 9), t3_count=(0, 0, 0, 1, 2),
        pre_seed=700_000, seed=2_800_000, series_a=7_500_000,
    ),
    "health-patient-correspondence": Drivers(
        "Epistolē", "health", "ΕΣΥ + ΗΔΙΚΑ + ΕΟΠΥΥ",
        tier1_price=75_000, tier2_acv=300_000, tier3_acv=1_200_000,
        seats_per_t2=120, seat_price=2500,
        t1_count=(3, 3, 2, 1, 0), t2_count=(0, 2, 4, 7, 10), t3_count=(0, 0, 1, 1, 2),
        pre_seed=500_000, seed=2_500_000, series_a=8_000_000,
        gross_margin_target=0.80,
    ),
    "transport-kteo-anomaly-agent": Drivers(
        "KTEO-Sentinel", "infrastructure-transport", "Γ.Γ. Μεταφορών + ΓΓΠΣΨΔ",
        tier1_price=100_000, tier2_acv=1_500_000, tier3_acv=3_500_000,
        t1_count=(1, 2, 1, 0, 0), t2_count=(0, 1, 2, 3, 4), t3_count=(0, 0, 0, 1, 2),
        pre_seed=800_000, seed=3_500_000, series_a=10_000_000,
    ),
    "maritime-fueleu-psc-assistant": Drivers(
        "PSC Copilot Hellas", "maritime-insular", "Hellenic Coast Guard / ΥΝΑΝΠ",
        tier1_price=110_000, tier2_acv=320_000, tier3_acv=1_000_000,
        seats_per_t2=8, seat_price=40000,  # per-port
        t1_count=(3, 3, 1, 0, 0), t2_count=(0, 2, 5, 9, 13), t3_count=(0, 0, 0, 1, 2),
        pre_seed=400_000, seed=2_500_000, series_a=7_000_000,
    ),
    "migration-press-qa-draft": Drivers(
        "Alkyoni", "migration-asylum", "ΥΜΑ Press Office",
        tier1_price=45_000, tier2_acv=100_000, tier3_acv=400_000,
        t1_count=(1, 2, 1, 0, 0), t2_count=(0, 2, 4, 7, 10), t3_count=(0, 0, 1, 2, 3),
        pre_seed=400_000, seed=1_500_000, series_a=4_500_000,
        engineering=(1, 2, 3, 5, 7), gross_margin_target=0.82,
    ),
    "digital-gov-civil-servant-process-assistant": Drivers(
        "Mitos Co-Pilot", "digital-governance", "ΓΓΠΣΨΔ + cross-ministry",
        tier1_price=200_000, tier2_acv=600_000, tier3_acv=2_400_000,
        seats_per_t2=2000, seat_price=300,  # €25/seat/mo × 12
        t1_count=(2, 2, 1, 0, 0), t2_count=(0, 2, 5, 9, 14), t3_count=(0, 0, 1, 2, 4),
        pre_seed=900_000, seed=4_000_000, series_a=12_000_000,
        engineering=(2, 5, 8, 12, 16), gross_margin_target=0.74,
    ),
    "health-pre-authorisation-drafter": Drivers(
        "Prograde Health", "health", "EOPYY Governor",
        tier1_price=90_000, tier2_acv=375_000, tier3_acv=1_400_000,
        seats_per_t2=150, seat_price=2500,
        t1_count=(2, 3, 2, 0, 0), t2_count=(0, 2, 4, 7, 10), t3_count=(0, 0, 1, 1, 2),
        pre_seed=750_000, seed=3_200_000, series_a=9_000_000,
    ),
    "mfa-sister-cities-diplomacy": Drivers(
        "Φιλόξενος", "foreign-affairs", "ΥΠΕΞ ΓΓΑΕ + KEDE",
        tier1_price=60_000, tier2_acv=140_000, tier3_acv=500_000,
        t1_count=(2, 3, 2, 1, 0), t2_count=(0, 2, 5, 9, 14), t3_count=(0, 0, 1, 1, 2),
        pre_seed=350_000, seed=1_500_000, series_a=4_000_000,
        engineering=(1, 2, 3, 5, 7),
    ),
    "migration-residence-permit-triage": Drivers(
        "Permit Triage", "migration-asylum", "Γ.Γ. Μεταναστευτικής Πολιτικής",
        tier1_price=130_000, tier2_acv=600_000, tier3_acv=1_800_000,
        seats_per_t2=80, seat_price=4500,
        t1_count=(2, 2, 1, 0, 0), t2_count=(0, 1, 3, 5, 8), t3_count=(0, 0, 1, 1, 2),
        pre_seed=750_000, seed=3_300_000, series_a=9_000_000,
    ),
    "defence-procurement-clause-checker": Drivers(
        "Klafsi Defence", "national-defence", "ΓΔΑΕΕ / ΥΠΕΘΑ",
        tier1_price=90_000, tier2_acv=400_000, tier3_acv=1_500_000,
        t1_count=(2, 2, 1, 0, 0), t2_count=(0, 1, 3, 5, 7), t3_count=(0, 0, 1, 1, 2),
        pre_seed=800_000, seed=3_500_000, series_a=10_000_000,
    ),
    "agri-elga-claims-triage": Drivers(
        "Chersos Claims", "rural-development", "ELGA President + ΥΠΑΑΤ",
        tier1_price=55_000, tier2_acv=225_000, tier3_acv=1_000_000,
        t1_count=(2, 3, 2, 0, 0), t2_count=(0, 2, 4, 7, 10), t3_count=(0, 0, 1, 1, 2),
        pre_seed=600_000, seed=2_500_000, series_a=7_000_000,
    ),
    "environment-eia-drafting-agent": Drivers(
        "EIA Drafter", "environment-energy", "ΥΠΕΝ ΔΙΠΑ + consultancies",
        tier1_price=150_000, tier2_acv=400_000, tier3_acv=1_200_000,
        seats_per_t2=60, seat_price=4800,
        t1_count=(1, 2, 1, 0, 0), t2_count=(0, 1, 3, 6, 10), t3_count=(0, 0, 0, 1, 2),
        pre_seed=650_000, seed=2_800_000, series_a=8_000_000,
    ),
    "health-chronic-prescription-review": Drivers(
        "Polyfarmaco Copilot", "health", "ΙΔΙΚΑ + GS Public Health + ΕΟΠΥΥ",
        tier1_price=150_000, tier2_acv=420_000, tier3_acv=1_500_000,
        seats_per_t2=140, seat_price=3000,
        t1_count=(1, 2, 1, 0, 0), t2_count=(0, 1, 3, 5, 8), t3_count=(0, 0, 1, 1, 2),
        pre_seed=900_000, seed=4_000_000, series_a=12_000_000,
        engineering=(2, 4, 7, 10, 13), gross_margin_target=0.74,
    ),
    "pmo-council-of-ministers-prep": Drivers(
        "Praktiko", "state-cross-cutting", "ΓΓΠΘ (PM Office)",
        tier1_price=100_000, tier2_acv=550_000, tier3_acv=1_800_000,
        t1_count=(2, 2, 1, 0, 0), t2_count=(0, 1, 3, 5, 7), t3_count=(0, 0, 1, 2, 3),
        pre_seed=800_000, seed=3_500_000, series_a=9_500_000,
        gross_margin_target=0.78,
    ),
    "education-school-inspection-drafting": Drivers(
        "Sýmvoulos Copilot", "education", "ΥΠΑΙΘ Γ.Γ. Π/Δ + ΙΕΠ",
        tier1_price=50_000, tier2_acv=250_000, tier3_acv=900_000,
        seats_per_t2=150, seat_price=1700,
        t1_count=(2, 3, 2, 1, 0), t2_count=(0, 1, 3, 5, 8), t3_count=(0, 0, 1, 1, 2),
        pre_seed=500_000, seed=2_200_000, series_a=6_000_000,
    ),
    "sepe-inspection-prioritiser": Drivers(
        "Ergani Risk Lens", "labour-social-security", "ΣΕΠΕ / ΥΠΕΚΥΠ",
        tier1_price=48_000, tier2_acv=300_000, tier3_acv=1_100_000,
        t1_count=(2, 3, 2, 0, 0), t2_count=(0, 2, 4, 7, 10), t3_count=(0, 0, 0, 1, 2),
        pre_seed=600_000, seed=2_600_000, series_a=7_500_000,
    ),
    "efka-appeal-response-drafter": Drivers(
        "Antikrousi", "labour-social-security", "ΝΣΚ + EFKA Governor",
        tier1_price=70_000, tier2_acv=320_000, tier3_acv=1_200_000,
        seats_per_t2=60, seat_price=5300,
        t1_count=(1, 2, 1, 0, 0), t2_count=(0, 1, 3, 5, 7), t3_count=(0, 0, 0, 1, 2),
        pre_seed=500_000, seed=2_500_000, series_a=7_000_000,
        engineering=(1, 3, 5, 7, 10),
    ),
}


# ---- Excel builder --------------------------------------------------------


def _header(ws, text):
    ws.append([text])
    cell = ws.cell(row=ws.max_row, column=1)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[ws.max_row].height = 22


def _section(ws, text):
    ws.append([text])
    cell = ws.cell(row=ws.max_row, column=1)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL


def _set_year_header(ws, row, start_col=2, label_col_text="Driver"):
    ws.cell(row=row, column=1, value=label_col_text).font = BOLD
    for i, y in enumerate(["Y1", "Y2", "Y3", "Y4", "Y5"], start=start_col):
        c = ws.cell(row=row, column=i, value=y)
        c.font = BOLD
        c.alignment = Alignment(horizontal="center")
        c.fill = SECTION_FILL


def _autofit(ws, default=14, first=44):
    ws.column_dimensions["A"].width = first
    for i in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(i)].width = default


def build_assumptions(ws, d: Drivers):
    """Returns dict mapping driver name → cell ref (e.g. 'Assumptions!B5')."""
    refs: dict[str, str] = {}
    _header(ws, f"Assumptions — {d.company}")
    ws.append([])
    ws.append([f"Ministry: {d.ministry}    Buyer: {d.buyer}"])
    ws.cell(row=ws.max_row, column=1).font = ITAL
    ws.append([])

    def row(label, vals, fill=INPUT_FILL, num=EUR, bold=False, key=None):
        ws.append([label] + list(vals))
        r = ws.max_row
        ws.cell(row=r, column=1).font = BOLD if bold else Font()
        for i, _ in enumerate(vals, start=2):
            c = ws.cell(row=r, column=i)
            c.fill = fill
            c.number_format = num
            if bold:
                c.font = BOLD
        if key:
            if isinstance(vals, (list, tuple)) and len(vals) == 5:
                refs[key] = [f"Assumptions!{get_column_letter(i)}{r}" for i in range(2, 7)]
            else:
                refs[key] = f"Assumptions!B{r}"

    _section(ws, "PRICING")
    _set_year_header(ws, ws.max_row + 1, label_col_text="Pricing")
    row("Tier 1 — pilot price (€)", [d.tier1_price]*5, num=EUR, key="tier1_price_arr")
    row("Tier 2 — year-one ACV (€)", [d.tier2_acv]*5, num=EUR, key="tier2_acv_arr")
    row("Tier 3 — framework ACV (€)", [d.tier3_acv]*5, num=EUR, key="tier3_acv_arr")
    row("Avg seats per Tier-2 customer", [d.seats_per_t2]*5, num=INT, key="seats_per_t2_arr")
    row("Blended seat price (€/yr)", [d.seat_price]*5, num=EUR, key="seat_price_arr")
    ws.append([])

    _section(ws, "ADOPTION (count per year)")
    _set_year_header(ws, ws.max_row + 1, label_col_text="Adoption")
    row("Tier 1 customers (new + active)", list(d.t1_count), num=INT, key="t1_count")
    row("Tier 2 customers (cumulative active)", list(d.t2_count), num=INT, key="t2_count")
    row("Tier 3 customers (cumulative active)", list(d.t3_count), num=INT, key="t3_count")
    ws.append([])

    _section(ws, "HEADCOUNT (FTE per year)")
    _set_year_header(ws, ws.max_row + 1, label_col_text="Role")
    row("Founders", list(d.founders), num=INT, key="founders")
    row("Engineering", list(d.engineering), num=INT, key="engineering")
    row("Product", list(d.product), num=INT, key="product")
    row("Sales", list(d.sales), num=INT, key="sales")
    row("Customer Success", list(d.cs), num=INT, key="cs")
    row("G&A", list(d.ga), num=INT, key="ga")
    row("Loaded €/FTE (Athens)", [d.fte_cost]*5, num=EUR, key="fte_cost_arr")
    ws.append([])

    _section(ws, "COGS & OPEX RATIOS")
    _set_year_header(ws, ws.max_row + 1, label_col_text="Cost driver")
    row("Hosting €/customer/yr", [d.hosting_per_customer]*5, num=EUR, key="hosting_arr")
    row("S&M non-personnel (% of revenue)", [d.sm_pct]*5, num=PCT, key="sm_pct_arr")
    row("G&A non-personnel (% of revenue)", [d.ga_pct]*5, num=PCT, key="ga_pct_arr")
    row("Customer-success additional COGS (% of revenue)", [d.cs_pct_revenue]*5, num=PCT, key="cs_cogs_pct_arr")
    row("Gross margin target", [d.gross_margin_target]*5, num=PCT, key="gm_target_arr")
    ws.append([])

    _section(ws, "CAPITAL")
    ws.append(["Pre-seed (€)", d.pre_seed])
    refs["pre_seed"] = f"Assumptions!B{ws.max_row}"
    ws.cell(row=ws.max_row, column=2).fill = INPUT_FILL
    ws.cell(row=ws.max_row, column=2).number_format = EUR
    ws.append(["Pre-seed raise year", d.pre_seed_year])
    refs["pre_seed_year"] = f"Assumptions!B{ws.max_row}"
    ws.cell(row=ws.max_row, column=2).fill = INPUT_FILL
    ws.append(["Seed (€)", d.seed])
    refs["seed"] = f"Assumptions!B{ws.max_row}"
    ws.cell(row=ws.max_row, column=2).fill = INPUT_FILL
    ws.cell(row=ws.max_row, column=2).number_format = EUR
    ws.append(["Seed raise year", d.seed_year])
    refs["seed_year"] = f"Assumptions!B{ws.max_row}"
    ws.cell(row=ws.max_row, column=2).fill = INPUT_FILL
    ws.append(["Series A (€)", d.series_a])
    refs["series_a"] = f"Assumptions!B{ws.max_row}"
    ws.cell(row=ws.max_row, column=2).fill = INPUT_FILL
    ws.cell(row=ws.max_row, column=2).number_format = EUR
    ws.append(["Series A raise year", d.series_a_year])
    refs["series_a_year"] = f"Assumptions!B{ws.max_row}"
    ws.cell(row=ws.max_row, column=2).fill = INPUT_FILL

    ws.freeze_panes = "B2"
    _autofit(ws)
    return refs


def build_revenue(ws, refs):
    _header(ws, "Revenue Build")
    ws.append([])
    _set_year_header(ws, ws.max_row + 1, label_col_text="Revenue line")
    # Tier 1
    r = ws.max_row + 1
    ws.cell(row=r, column=1, value="Tier 1 — pilots × price (€)").font = BOLD
    for i, col in enumerate(["B", "C", "D", "E", "F"]):
        ws.cell(row=r, column=2+i, value=f"={refs['t1_count'][i]}*{refs['tier1_price_arr'][i]}").number_format = EUR
    r1_row = r
    # Tier 2 (subscription ACV)
    ws.append(["Tier 2 — customers × ACV (€)"] + [None]*5)
    r = ws.max_row
    ws.cell(row=r, column=1).font = BOLD
    for i in range(5):
        ws.cell(row=r, column=2+i, value=f"={refs['t2_count'][i]}*{refs['tier2_acv_arr'][i]}").number_format = EUR
    r2_row = r
    # Tier 3
    ws.append(["Tier 3 — frameworks × ACV (€)"] + [None]*5)
    r = ws.max_row
    ws.cell(row=r, column=1).font = BOLD
    for i in range(5):
        ws.cell(row=r, column=2+i, value=f"={refs['t3_count'][i]}*{refs['tier3_acv_arr'][i]}").number_format = EUR
    r3_row = r
    # Total
    ws.append([])
    ws.append(["TOTAL REVENUE (€)"] + [None]*5)
    rt = ws.max_row
    ws.cell(row=rt, column=1).font = BOLD
    for i, col in enumerate(["B", "C", "D", "E", "F"]):
        f = f"={col}{r1_row}+{col}{r2_row}+{col}{r3_row}"
        c = ws.cell(row=rt, column=2+i, value=f)
        c.fill = TOTAL_FILL
        c.font = BOLD
        c.number_format = EUR
    total_row = rt
    # Net new ARR
    ws.append(["Net new revenue (€)", None] + [None]*4)
    rn = ws.max_row
    ws.cell(row=rn, column=1).font = BOLD
    ws.cell(row=rn, column=2, value="").number_format = EUR
    for i, col in enumerate(["C", "D", "E", "F"], start=1):
        prev = ["B", "C", "D", "E"][i-1]
        c = ws.cell(row=rn, column=2+i, value=f"={col}{total_row}-{prev}{total_row}")
        c.number_format = EUR

    refs["revenue_row"] = total_row
    ws.freeze_panes = "B2"
    _autofit(ws)


def build_headcount_costs(ws, refs):
    _header(ws, "Headcount & Costs")
    ws.append([])
    _set_year_header(ws, ws.max_row + 1, label_col_text="Cost line")
    # Total FTE
    r = ws.max_row + 1
    ws.cell(row=r, column=1, value="Total FTE").font = BOLD
    for i in range(5):
        f = f"={refs['founders'][i]}+{refs['engineering'][i]}+{refs['product'][i]}+{refs['sales'][i]}+{refs['cs'][i]}+{refs['ga'][i]}"
        ws.cell(row=r, column=2+i, value=f).number_format = INT
    fte_row = r
    # People cost
    ws.append(["People cost (€)"] + [None]*5)
    r = ws.max_row
    ws.cell(row=r, column=1).font = BOLD
    for i, col in enumerate(["B", "C", "D", "E", "F"]):
        f = f"={col}{fte_row}*{refs['fte_cost_arr'][i]}"
        ws.cell(row=r, column=2+i, value=f).number_format = EUR
    people_row = r
    # Hosting
    ws.append(["Hosting cost (€)"] + [None]*5)
    r = ws.max_row
    ws.cell(row=r, column=1).font = BOLD
    for i in range(5):
        n_customers = f"({refs['t1_count'][i]}+{refs['t2_count'][i]}+{refs['t3_count'][i]})"
        f = f"={n_customers}*{refs['hosting_arr'][i]}"
        ws.cell(row=r, column=2+i, value=f).number_format = EUR
    hosting_row = r
    # S&M non-personnel
    ws.append(["S&M non-personnel (€)"] + [None]*5)
    r = ws.max_row
    ws.cell(row=r, column=1).font = BOLD
    rev_row = refs["revenue_row"]
    for i, col in enumerate(["B", "C", "D", "E", "F"]):
        f = f"=Revenue!{col}{rev_row}*{refs['sm_pct_arr'][i]}"
        ws.cell(row=r, column=2+i, value=f).number_format = EUR
    sm_row = r
    # G&A non-personnel
    ws.append(["G&A non-personnel (€)"] + [None]*5)
    r = ws.max_row
    ws.cell(row=r, column=1).font = BOLD
    for i, col in enumerate(["B", "C", "D", "E", "F"]):
        f = f"=Revenue!{col}{rev_row}*{refs['ga_pct_arr'][i]}"
        ws.cell(row=r, column=2+i, value=f).number_format = EUR
    ga_row = r
    # CS additional COGS
    ws.append(["CS additional COGS (€)"] + [None]*5)
    r = ws.max_row
    ws.cell(row=r, column=1).font = BOLD
    for i, col in enumerate(["B", "C", "D", "E", "F"]):
        f = f"=Revenue!{col}{rev_row}*{refs['cs_cogs_pct_arr'][i]}"
        ws.cell(row=r, column=2+i, value=f).number_format = EUR
    cs_cogs_row = r
    ws.append([])
    ws.append(["TOTAL OPERATING COST (€)"] + [None]*5)
    rt = ws.max_row
    ws.cell(row=rt, column=1).font = BOLD
    for i, col in enumerate(["B", "C", "D", "E", "F"]):
        f = f"={col}{people_row}+{col}{hosting_row}+{col}{sm_row}+{col}{ga_row}+{col}{cs_cogs_row}"
        c = ws.cell(row=rt, column=2+i, value=f)
        c.fill = TOTAL_FILL
        c.font = BOLD
        c.number_format = EUR
    refs["fte_row"] = fte_row
    refs["people_row"] = people_row
    refs["hosting_row"] = hosting_row
    refs["sm_row"] = sm_row
    refs["ga_row"] = ga_row
    refs["cs_cogs_row"] = cs_cogs_row
    refs["total_cost_row"] = rt

    ws.freeze_panes = "B2"
    _autofit(ws)


def build_pnl(ws, refs):
    _header(ws, "P&L (5 years)")
    ws.append([])
    _set_year_header(ws, ws.max_row + 1, label_col_text="P&L line")
    rev_row = refs["revenue_row"]
    cols = ["B", "C", "D", "E", "F"]
    # Revenue
    r = ws.max_row + 1
    ws.cell(row=r, column=1, value="Revenue").font = BOLD
    for i, col in enumerate(cols):
        ws.cell(row=r, column=2+i, value=f"=Revenue!{col}{rev_row}").number_format = EUR
    pl_rev = r
    # COGS = hosting + cs_cogs
    ws.append(["COGS"] + [None]*5)
    r = ws.max_row
    for i, col in enumerate(cols):
        f = f"=Headcount_Costs!{col}{refs['hosting_row']}+Headcount_Costs!{col}{refs['cs_cogs_row']}"
        ws.cell(row=r, column=2+i, value=f).number_format = EUR
    pl_cogs = r
    # Gross profit
    ws.append(["Gross Profit"] + [None]*5)
    r = ws.max_row
    ws.cell(row=r, column=1).font = BOLD
    for i, col in enumerate(cols):
        f = f"={col}{pl_rev}-{col}{pl_cogs}"
        c = ws.cell(row=r, column=2+i, value=f)
        c.font = BOLD
        c.number_format = EUR
    pl_gp = r
    # Gross margin
    ws.append(["Gross Margin %"] + [None]*5)
    r = ws.max_row
    for i, col in enumerate(cols):
        f = f"=IFERROR({col}{pl_gp}/{col}{pl_rev},0)"
        c = ws.cell(row=r, column=2+i, value=f)
        c.number_format = PCT
    pl_gm = r
    ws.append([])
    # Opex
    ws.append(["Personnel (R&D + S&M + G&A + CS)"] + [None]*5)
    r = ws.max_row
    for i, col in enumerate(cols):
        f = f"=Headcount_Costs!{col}{refs['people_row']}"
        ws.cell(row=r, column=2+i, value=f).number_format = EUR
    pl_people = r
    ws.append(["S&M non-personnel"] + [None]*5)
    r = ws.max_row
    for i, col in enumerate(cols):
        f = f"=Headcount_Costs!{col}{refs['sm_row']}"
        ws.cell(row=r, column=2+i, value=f).number_format = EUR
    pl_sm = r
    ws.append(["G&A non-personnel"] + [None]*5)
    r = ws.max_row
    for i, col in enumerate(cols):
        f = f"=Headcount_Costs!{col}{refs['ga_row']}"
        ws.cell(row=r, column=2+i, value=f).number_format = EUR
    pl_ga = r
    ws.append(["Total OPEX"] + [None]*5)
    r = ws.max_row
    ws.cell(row=r, column=1).font = BOLD
    for i, col in enumerate(cols):
        f = f"={col}{pl_people}+{col}{pl_sm}+{col}{pl_ga}"
        c = ws.cell(row=r, column=2+i, value=f)
        c.font = BOLD
        c.number_format = EUR
    pl_opex = r
    ws.append([])
    # EBITDA
    ws.append(["EBITDA"] + [None]*5)
    r = ws.max_row
    ws.cell(row=r, column=1).font = BOLD
    for i, col in enumerate(cols):
        f = f"={col}{pl_gp}-{col}{pl_opex}"
        c = ws.cell(row=r, column=2+i, value=f)
        c.fill = TOTAL_FILL
        c.font = BOLD
        c.number_format = EUR
    pl_ebitda = r
    ws.append(["EBITDA margin %"] + [None]*5)
    r = ws.max_row
    for i, col in enumerate(cols):
        f = f"=IFERROR({col}{pl_ebitda}/{col}{pl_rev},0)"
        c = ws.cell(row=r, column=2+i, value=f)
        c.number_format = PCT

    refs["pl_rev"] = pl_rev
    refs["pl_ebitda"] = pl_ebitda
    refs["pl_gm"] = pl_gm
    refs["pl_gp"] = pl_gp

    ws.freeze_panes = "B2"
    _autofit(ws)


def build_cashflow(ws, refs):
    _header(ws, "Cash Flow & Runway")
    ws.append([])
    _set_year_header(ws, ws.max_row + 1, label_col_text="Cash line")
    cols = ["B", "C", "D", "E", "F"]
    # Equity raised — formula references the year-mapped raise
    ws.append(["Equity raised (€)"] + [None]*5)
    r = ws.max_row
    for i, col in enumerate(cols):
        year = i + 1
        f = (f"=IF({refs['pre_seed_year']}={year},{refs['pre_seed']},0)"
             f"+IF({refs['seed_year']}={year},{refs['seed']},0)"
             f"+IF({refs['series_a_year']}={year},{refs['series_a']},0)")
        ws.cell(row=r, column=2+i, value=f).number_format = EUR
    cf_eq = r
    # EBITDA flow-through
    ws.append(["EBITDA"] + [None]*5)
    r = ws.max_row
    for i, col in enumerate(cols):
        ws.cell(row=r, column=2+i, value=f"=P_and_L!{col}{refs['pl_ebitda']}").number_format = EUR
    cf_ebitda = r
    # WC change ~10% revenue
    ws.append(["Working capital change (10% Δrev)"] + [None]*5)
    r = ws.max_row
    rev_row = refs["revenue_row"]
    for i, col in enumerate(cols):
        if i == 0:
            f = f"=-Revenue!{col}{rev_row}*0.1"
        else:
            prev = cols[i-1]
            f = f"=-(Revenue!{col}{rev_row}-Revenue!{prev}{rev_row})*0.1"
        ws.cell(row=r, column=2+i, value=f).number_format = EUR
    cf_wc = r
    ws.append([])
    # Net cash flow
    ws.append(["Net cash flow"] + [None]*5)
    r = ws.max_row
    ws.cell(row=r, column=1).font = BOLD
    for i, col in enumerate(cols):
        f = f"={col}{cf_eq}+{col}{cf_ebitda}+{col}{cf_wc}"
        c = ws.cell(row=r, column=2+i, value=f)
        c.font = BOLD
        c.number_format = EUR
    cf_net = r
    # Starting / Ending cash
    ws.append(["Starting cash"] + [None]*5)
    r = ws.max_row
    for i, col in enumerate(cols):
        if i == 0:
            ws.cell(row=r, column=2+i, value=0).number_format = EUR
        else:
            prev = cols[i-1]
            ws.cell(row=r, column=2+i, value=f"={prev}{r+1}").number_format = EUR
    cf_start = r
    ws.append(["Ending cash"] + [None]*5)
    r = ws.max_row
    ws.cell(row=r, column=1).font = BOLD
    for i, col in enumerate(cols):
        f = f"={col}{cf_start}+{col}{cf_net}"
        c = ws.cell(row=r, column=2+i, value=f)
        c.fill = TOTAL_FILL
        c.font = BOLD
        c.number_format = EUR
    cf_end = r
    # Runway months
    ws.append(["Runway (months) @ year-end"] + [None]*5)
    r = ws.max_row
    for i, col in enumerate(cols):
        # monthly burn = -EBITDA/12 if EBITDA<0 else 0
        f = f"=IFERROR(IF({col}{cf_ebitda}<0,{col}{cf_end}/(-{col}{cf_ebitda}/12),999),0)"
        c = ws.cell(row=r, column=2+i, value=f)
        c.number_format = "#,##0"

    refs["cf_end"] = cf_end
    refs["cf_ebitda"] = cf_ebitda
    ws.freeze_panes = "B2"
    _autofit(ws)


def build_kpis(ws, refs, d: Drivers):
    _header(ws, "KPIs dashboard")
    ws.append([])
    _set_year_header(ws, ws.max_row + 1, label_col_text="KPI")
    cols = ["B", "C", "D", "E", "F"]
    # ARR (= revenue this year)
    rev_row = refs["revenue_row"]
    ws.append(["ARR end of year (€)"] + [None]*5)
    r = ws.max_row
    ws.cell(row=r, column=1).font = BOLD
    for i, col in enumerate(cols):
        c = ws.cell(row=r, column=2+i, value=f"=Revenue!{col}{rev_row}")
        c.fill = TOTAL_FILL
        c.font = BOLD
        c.number_format = EUR
    # Net new ARR
    ws.append(["Net new ARR (€)"] + [None]*5)
    r = ws.max_row
    for i, col in enumerate(cols):
        if i == 0:
            ws.cell(row=r, column=2+i, value=f"=Revenue!{col}{rev_row}").number_format = EUR
        else:
            prev = cols[i-1]
            ws.cell(row=r, column=2+i, value=f"=Revenue!{col}{rev_row}-Revenue!{prev}{rev_row}").number_format = EUR
    # Gross margin
    ws.append(["Gross margin %"] + [None]*5)
    r = ws.max_row
    for i, col in enumerate(cols):
        c = ws.cell(row=r, column=2+i, value=f"=P_and_L!{col}{refs['pl_gm']}")
        c.number_format = PCT
    # Burn / month
    ws.append(["Monthly burn (€)"] + [None]*5)
    r = ws.max_row
    for i, col in enumerate(cols):
        f = f"=IF(P_and_L!{col}{refs['pl_ebitda']}<0,-P_and_L!{col}{refs['pl_ebitda']}/12,0)"
        ws.cell(row=r, column=2+i, value=f).number_format = EUR
    # Customers
    ws.append(["Active customers"] + [None]*5)
    r = ws.max_row
    for i, col in enumerate(cols):
        f = f"={refs['t1_count'][i]}+{refs['t2_count'][i]}+{refs['t3_count'][i]}"
        ws.cell(row=r, column=2+i, value=f).number_format = INT
    # Headcount
    ws.append(["Total headcount"] + [None]*5)
    r = ws.max_row
    for i, col in enumerate(cols):
        ws.cell(row=r, column=2+i, value=f"=Headcount_Costs!{col}{refs['fte_row']}").number_format = INT
    # Revenue / FTE
    ws.append(["Revenue per FTE (€)"] + [None]*5)
    r = ws.max_row
    fte_row = refs["fte_row"]
    for i, col in enumerate(cols):
        f = f"=IFERROR(Revenue!{col}{rev_row}/Headcount_Costs!{col}{fte_row},0)"
        ws.cell(row=r, column=2+i, value=f).number_format = EUR
    # Cash at year-end
    ws.append(["Cash at year-end (€)"] + [None]*5)
    r = ws.max_row
    for i, col in enumerate(cols):
        ws.cell(row=r, column=2+i, value=f"=Cash_Flow!{col}{refs['cf_end']}").number_format = EUR
    ws.append([])
    # Scenario table
    _section(ws, "Y5 ARR scenarios (sensitivity on Tier 2/3 counts)")
    ws.append(["Scenario", "Multiplier", "Y5 ARR (€)"])
    for c in ws[ws.max_row]:
        c.font = BOLD
    ws.append(["Conservative", 0.6, f"=Revenue!F{rev_row}*0.6"])
    ws.cell(row=ws.max_row, column=3).number_format = EUR
    ws.append(["Base", 1.0, f"=Revenue!F{rev_row}"])
    ws.cell(row=ws.max_row, column=3).number_format = EUR
    ws.cell(row=ws.max_row, column=3).font = BOLD
    ws.cell(row=ws.max_row, column=3).fill = TOTAL_FILL
    ws.append(["Aggressive", 1.5, f"=Revenue!F{rev_row}*1.5"])
    ws.cell(row=ws.max_row, column=3).number_format = EUR

    ws.freeze_panes = "B2"
    _autofit(ws)


def build_model(idea_id: str, d: Drivers):
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    refs: dict = {}
    refs.update(build_assumptions(wb.create_sheet("Assumptions"), d) or {})
    # Need refs returned from assumptions
    ws_a = wb["Assumptions"]
    # We must rebuild refs because build_assumptions returns them
    # — but we already overwrote them. Fix: rerun explicitly:
    wb.remove(ws_a)
    ws_a = wb.create_sheet("Assumptions")
    refs = build_assumptions(ws_a, d)
    build_revenue(wb.create_sheet("Revenue"), refs)
    build_headcount_costs(wb.create_sheet("Headcount_Costs"), refs)
    build_pnl(wb.create_sheet("P_and_L"), refs)
    build_cashflow(wb.create_sheet("Cash_Flow"), refs)
    build_kpis(wb.create_sheet("KPIs"), refs, d)
    out = os.path.join(OUT_DIR, f"{idea_id}.xlsx")
    wb.save(out)
    size = os.path.getsize(out)
    return out, size


def main():
    results = []
    for idea_id, d in DRIVERS.items():
        try:
            path, size = build_model(idea_id, d)
            results.append((idea_id, path, size, None))
            print(f"OK {idea_id} {size//1024}KB")
        except Exception as e:
            results.append((idea_id, None, 0, str(e)))
            print(f"FAIL {idea_id}: {e}")
    print(f"\nTotal: {sum(1 for r in results if r[3] is None)}/{len(results)} built")
    return results


if __name__ == "__main__":
    main()
